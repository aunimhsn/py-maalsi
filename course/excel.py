import openpyxl

wb:openpyxl.Workbook = openpyxl.load_workbook('./data/transactions.xlsx')
ws = wb['sheet_1']

for i in range(2, ws.max_row + 1):
    float_price = float(ws[f'C{i}'].value.replace('€', '').replace(',', '.'))
    discounted_price = round(float_price * 0.9, 2)
    ws[f'D{i}'] = f'{discounted_price}€'

wb.save('./data/transactions_updated.xlsx')